import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter


# ── Colours ────────────────────────────────────────────────
BLUE_DARK   = "1A56DB"
BLUE_LIGHT  = "EFF6FF"
GREEN_DARK  = "059669"
GREEN_LIGHT = "ECFDF5"
AMBER       = "D97706"
AMBER_LIGHT = "FFFBEB"
GRAY_HDR    = "F3F4F6"
WHITE       = "FFFFFF"
BORDER_CLR  = "DEE2E6"

def _border():
    s = Side(style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000"):
    return Font(bold=bold, size=size, color=color)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def generate_excel(consultations, output_folder):
    """
    Generate an Excel report with:
    - Sheet 1: All consultations data
    - Sheet 2: Entity summary per consultation
    - Sheet 3: Charts (bar + pie)
    """
    os.makedirs(output_folder, exist_ok=True)
    filename = f"MediNote_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(output_folder, filename)

    wb = Workbook()

    _build_consultations_sheet(wb, consultations)
    _build_entities_sheet(wb, consultations)
    _build_charts_sheet(wb, consultations)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)
    print(f"Excel report generated: {filepath}")
    return filename


def _build_consultations_sheet(wb, consultations):
    ws = wb.create_sheet("Consultations")
    ws.sheet_view.showGridLines = False

    # Title row
    ws.merge_cells("A1:I1")
    ws["A1"] = "MediNote AI — Consultation Report"
    ws["A1"].font      = _font(bold=True, size=14, color=WHITE)
    ws["A1"].fill      = _hdr_fill(BLUE_DARK)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 32

    # Sub-title
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generated on {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A2"].font      = _font(size=9, color="6B7280")
    ws["A2"].alignment = _center()
    ws["A2"].fill      = _hdr_fill("F9FAFB")

    # Header row
    headers = ["#", "Report ID", "Patient Name", "Doctor",
               "Department", "Date", "Status", "Visit Type", "Chief Complaint"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = _font(bold=True, size=9, color=WHITE)
        cell.fill      = _hdr_fill(BLUE_DARK)
        cell.alignment = _center()
        cell.border    = _border()
    ws.row_dimensions[3].height = 22

    # Data rows
    for i, c in enumerate(consultations, 1):
        row = i + 3
        # Get department and visit type from entities
        dept    = _get_entity(c, "department")
        vtype   = _get_entity(c, "visit_type")
        chief   = _get_entity(c, "chief_complaint")

        values = [
            i,
            f"MN-{c.id:04d}",
            c.patient_name,
            c.doctor_name,
            dept or "General Medicine",
            c.created_at.strftime("%d %b %Y %H:%M"),
            c.status.upper(),
            vtype or "First visit",
            chief or "—"
        ]
        fill_color = GREEN_LIGHT if c.status == "approved" else \
                     AMBER_LIGHT if c.status == "ready"    else GRAY_HDR

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = _font(size=9)
            cell.fill      = _hdr_fill(fill_color)
            cell.alignment = _center() if col in [1,2,6,7,8] else _left()
            cell.border    = _border()

        ws.row_dimensions[row].height = 18

    # Column widths
    widths = [5, 10, 20, 18, 18, 18, 12, 14, 35]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_entities_sheet(wb, consultations):
    ws = wb.create_sheet("Clinical Entities")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "Clinical Entities per Consultation"
    ws["A1"].font      = _font(bold=True, size=13, color=WHITE)
    ws["A1"].fill      = _hdr_fill(GREEN_DARK)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 30

    headers = ["Report ID", "Patient", "Symptoms",
               "Medications", "Diagnoses", "Allergies", "Vitals", "Follow-up"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = _font(bold=True, size=9, color=WHITE)
        cell.fill      = _hdr_fill(GREEN_DARK)
        cell.alignment = _center()
        cell.border    = _border()
    ws.row_dimensions[2].height = 20

    entity_keys = ["symptoms","medications","diagnoses","allergies","vitals","follow_up"]
    alt = False
    for i, c in enumerate(consultations, 1):
        row   = i + 2
        alt   = not alt
        fcolor = "F9FAFB" if alt else WHITE

        row_data = [f"MN-{c.id:04d}", c.patient_name]
        for key in entity_keys:
            vals = [e.entity_value for e in c.entities if e.entity_type == key]
            row_data.append(", ".join(vals) if vals else "—")

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = _font(size=9)
            cell.fill      = _hdr_fill(fcolor)
            cell.alignment = _left()
            cell.border    = _border()
        ws.row_dimensions[row].height = 20

    widths = [10, 20, 30, 25, 25, 20, 20, 30]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_charts_sheet(wb, consultations):
    ws = wb.create_sheet("Analytics")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    ws["A1"] = "MediNote AI — Analytics Dashboard"
    ws["A1"].font      = _font(bold=True, size=13, color=WHITE)
    ws["A1"].fill      = _hdr_fill(BLUE_DARK)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 30

    # ── Status breakdown data ──────────────────────────────
    statuses = {}
    for c in consultations:
        statuses[c.status] = statuses.get(c.status, 0) + 1

    ws["A3"] = "Status"
    ws["B3"] = "Count"
    ws["A3"].font = ws["B3"].font = _font(bold=True, size=9, color=WHITE)
    ws["A3"].fill = ws["B3"].fill = _hdr_fill(BLUE_DARK)
    ws["A3"].alignment = ws["B3"].alignment = _center()

    status_labels = ["uploaded","processing","transcribed","ready","approved","error"]
    for r, label in enumerate(status_labels, 4):
        ws.cell(row=r, column=1, value=label.capitalize()).font = _font(size=9)
        ws.cell(row=r, column=2, value=statuses.get(label, 0)).font = _font(size=9)
        ws.cell(row=r, column=1).alignment = _left()
        ws.cell(row=r, column=2).alignment = _center()

    # ── Bar chart: consultations by status ─────────────────
    bar = BarChart()
    bar.type    = "col"
    bar.title   = "Consultations by Status"
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Status"
    bar.style   = 10
    bar.width   = 14
    bar.height  = 10

    data = Reference(ws, min_col=2, min_row=3, max_row=3+len(status_labels))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3+len(status_labels))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = BLUE_DARK
    ws.add_chart(bar, "D3")

    # ── Pie chart: approved vs pending ─────────────────────
    ws["A12"] = "Category"
    ws["B12"] = "Count"
    ws["A12"].font = ws["B12"].font = _font(bold=True, size=9, color=WHITE)
    ws["A12"].fill = ws["B12"].fill = _hdr_fill(GREEN_DARK)
    ws.cell(row=13, column=1, value="Approved").font  = _font(size=9)
    ws.cell(row=13, column=2, value=statuses.get("approved",0)).font = _font(size=9)
    ws.cell(row=14, column=1, value="Pending").font   = _font(size=9)
    ws.cell(row=14, column=2,
            value=len(consultations)-statuses.get("approved",0)).font = _font(size=9)

    pie = PieChart()
    pie.title  = "Approved vs Pending"
    pie.style  = 10
    pie.width  = 14
    pie.height = 10

    p_data = Reference(ws, min_col=2, min_row=12, max_row=14)
    p_cats = Reference(ws, min_col=1, min_row=13, max_row=14)
    pie.add_data(p_data, titles_from_data=True)
    pie.set_categories(p_cats)

    # Colour slices
    slice_colors = [GREEN_DARK, AMBER]
    for idx, color in enumerate(slice_colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        pie.series[0].dPt.append(pt)

    ws.add_chart(pie, "D14")

    # ── Monthly trend data ─────────────────────────────────
    monthly = {}
    for c in consultations:
        key = c.created_at.strftime("%b %Y")
        monthly[key] = monthly.get(key, 0) + 1

    ws["A24"] = "Month"
    ws["B24"] = "Consultations"
    ws["A24"].font = ws["B24"].font = _font(bold=True, size=9, color=WHITE)
    ws["A24"].fill = ws["B24"].fill = _hdr_fill(AMBER)

    for r, (month, count) in enumerate(monthly.items(), 25):
        ws.cell(row=r, column=1, value=month).font  = _font(size=9)
        ws.cell(row=r, column=2, value=count).font  = _font(size=9)
        ws.cell(row=r, column=1).alignment = _left()
        ws.cell(row=r, column=2).alignment = _center()

    if monthly:
        trend = BarChart()
        trend.type   = "col"
        trend.title  = "Monthly Consultation Volume"
        trend.style  = 10
        trend.width  = 14
        trend.height = 10
        t_end = 24 + len(monthly)
        t_data = Reference(ws, min_col=2, min_row=24, max_row=t_end)
        t_cats = Reference(ws, min_col=1, min_row=25, max_row=t_end)
        trend.add_data(t_data, titles_from_data=True)
        trend.set_categories(t_cats)
        trend.series[0].graphicalProperties.solidFill = AMBER
        ws.add_chart(trend, "D24")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14


def _get_entity(consultation, entity_type):
    """Get first entity value of a given type from a consultation."""
    for e in consultation.entities:
        if e.entity_type == entity_type and e.entity_value:
            return e.entity_value
    return ""